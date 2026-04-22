"""从项目根目录 test-data/sample.xlsx 加载示例表格数据的路由，并提供通用文件导入接口。"""
from __future__ import annotations

import csv
import time
from datetime import datetime
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Iterable, List

from fastapi import APIRouter, File, HTTPException, UploadFile
from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict, Field

from .export import SchemaCol
from app.logging_config import get_logger
from app.services.projects import project_store

router = APIRouter(prefix="/api", tags=["load"])
logger = get_logger("api.load")


class LoadedTable(BaseModel):
    """单张从上传文件载入的表。"""

    model_config = ConfigDict(populate_by_name=True, serialize_by_alias=True)

    name: str
    rows: List[dict[str, Any]] = Field(default_factory=list)
    table_schema: List[SchemaCol] = Field(
        default_factory=list,
        validation_alias="schema",
        serialization_alias="schema",
    )


class LoadSampleResponse(BaseModel):
    """加载示例数据或导入文件的响应结构，包含后端创建的 projectId。"""

    projectId: str
    tables: List[LoadedTable] = Field(default_factory=list)


def _project_root() -> Path:
    """推导项目根目录（server 上一级）。

    优先通过在当前文件父目录中定位名为 "server" 的目录，以减少对层级深度的硬编码依赖。
    若未找到，则回退到历史实现（向上回溯四级目录）。
    """
    current = Path(__file__).resolve()
    for parent in current.parents:
        if parent.name == "server":
            return parent.parent
    # 回退策略：假设项目结构仍为 <project>/server/app/...
    # 若 parents 深度不足，将由 Python 抛出 IndexError，方便在开发阶段暴露问题。
    return current.parents[4]


def _test_data_excel_path() -> Path:
    """返回示例 Excel 的完整路径。"""
    return _project_root() / "test-data" / "sample.xlsx"


def _infer_type_from_values(values: Iterable[Any]) -> str:
    """根据一列中的值推断类型：number/string/date。"""
    for v in values:
        if v is None:
            continue
        if isinstance(v, (int, float)):
            return "number"
        if isinstance(v, datetime):
            return "date"
        # 兜底走 string
        return "string"
    return "string"


def _load_tables_from_excel_stream(
    data: bytes, *, max_rows_per_sheet: int | None = None
) -> list[LoadedTable]:
    """从 Excel 二进制流中解析所有 sheet 为 LoadedTable。

    Args:
        data: Excel 文件的二进制内容。
        max_rows_per_sheet: 每个工作表最多加载的行数上限；为 None 时不限制。

    Returns:
        解析得到的 LoadedTable 列表。
    """
    wb = load_workbook(BytesIO(data), data_only=True)
    tables: list[LoadedTable] = []

    for ws in wb.worksheets:
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            # 空 sheet，跳过
            continue
        if not header:
            continue

        keys = [str(k) if k is not None else "" for k in header]
        keys = [k for k in keys if k]  # 去掉空表头
        if not keys:
            continue

        raw_rows: list[tuple[Any, ...]] = []
        for r in rows_iter:
            if max_rows_per_sheet is not None and len(raw_rows) >= max_rows_per_sheet:
                break
            raw_rows.append(r)
        dict_rows: list[dict[str, Any]] = []
        for r in raw_rows:
            if r is None:
                continue
            # 对齐 keys 长度
            values = list(r)[: len(keys)]
            if all(v is None for v in values):
                continue
            dict_rows.append({k: values[i] for i, k in enumerate(keys)})

        # 推断 schema
        schema: list[SchemaCol] = []
        for col_idx, key in enumerate(keys):
            col_values = (
                row[col_idx] if len(row) > col_idx else None for row in raw_rows
            )
            t = _infer_type_from_values(col_values)
            schema.append(SchemaCol(key=key, type=t))

        tables.append(LoadedTable(name=ws.title, rows=dict_rows, table_schema=schema))

    return tables


def _load_tables_from_excel(
    path: Path, *, max_rows_per_sheet: int | None = None
) -> list[LoadedTable]:
    """从给定 Excel 文件路径解析所有 sheet 为 LoadedTable。

    Args:
        path: Excel 文件路径。
        max_rows_per_sheet: 每个工作表最多加载的行数上限；为 None 时不限制。

    Returns:
        解析得到的 LoadedTable 列表。
    """
    if not path.exists():
        raise HTTPException(status_code=404, detail=f"示例文件不存在: {path}")
    start = time.monotonic()
    data = path.read_bytes()
    tables = _load_tables_from_excel_stream(
        data, max_rows_per_sheet=max_rows_per_sheet
    )
    elapsed = time.monotonic() - start
    total_rows = sum(len(t.rows) for t in tables)
    total_cols = sum(len(t.table_schema) for t in tables)
    logger.info(
        "Loaded Excel file %s into %d tables (rows=%d, cols=%d, max_rows_per_sheet=%s) in %.3fs",
        path,
        len(tables),
        total_rows,
        total_cols,
        max_rows_per_sheet,
        elapsed,
    )
    return tables


def _load_tables_from_csv(
    name: str, data: bytes, *, max_rows: int | None = None
) -> list[LoadedTable]:
    """从 CSV 内容解析为单张 LoadedTable。

    默认认为首行为表头，其余行为数据行。
    """
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        # 兜底改用系统默认编码，错误由上层捕获。
        text = data.decode()

    start = time.monotonic()
    reader = csv.reader(StringIO(text))
    try:
        header = next(reader)
    except StopIteration:
        return []
    if not header:
        return []

    keys = [str(k).strip() for k in header if str(k).strip()]
    if not keys:
        return []

    raw_rows: list[list[Any]] = []
    dict_rows: list[dict[str, Any]] = []
    for row in reader:
        if max_rows is not None and len(raw_rows) >= max_rows:
            break
        if not row:
            continue
        values = row[: len(keys)]
        if all((v is None or str(v).strip() == "") for v in values):
            continue
        raw_rows.append(values)
        dict_rows.append(
            {k: values[i] if i < len(values) else None for i, k in enumerate(keys)}
        )

    schema: list[SchemaCol] = []
    for col_idx, key in enumerate(keys):
        col_values = (
            r[col_idx] if len(r) > col_idx else None for r in raw_rows
        )
        t = _infer_type_from_values(col_values)
        schema.append(SchemaCol(key=key, type=t))

    if not dict_rows:
        return []

    tables = [LoadedTable(name=name, rows=dict_rows, table_schema=schema)]
    elapsed = time.monotonic() - start
    total_rows = sum(len(t.rows) for t in tables)
    total_cols = sum(len(t.table_schema) for t in tables)
    logger.info(
        "Loaded CSV %s into %d tables (rows=%d, cols=%d, max_rows=%s) in %.3fs",
        name,
        len(tables),
        total_rows,
        total_cols,
        max_rows,
        elapsed,
    )
    return tables


@router.get("/load-sample", response_model=LoadSampleResponse)
async def load_sample() -> LoadSampleResponse:
    """从 test-data/sample.xlsx 加载示例表格数据，返回表名 + schema + rows。"""
    path = _test_data_excel_path()
    logger.info("load_sample start path=%s", path)
    tables = _load_tables_from_excel(path)
    if not tables:
        raise HTTPException(
            status_code=500,
            detail="示例文件中未找到任何表数据",
        )

    # 将加载的表注册到后端 ProjectStore 中，生成 projectId。
    tables_dict = {
        t.name: {
            "name": t.name,
            "rows": t.rows,
            "schema": [s.model_dump(mode="python") for s in t.table_schema],
        }
        for t in tables
    }
    state = project_store.create_project(tables_dict)
    logger.info(
        "load_sample done project_id=%s tables=%d",
        state.id,
        len(tables),
    )
    return LoadSampleResponse(projectId=state.id, tables=tables)


@router.post("/import-file", response_model=LoadSampleResponse)
async def import_file(file: UploadFile = File(...)) -> LoadSampleResponse:
    """从上传的 Excel/CSV 文件创建新的 ProjectState 并返回 projectId 与表列表。"""
    if not file.filename:
        raise HTTPException(status_code=400, detail="未提供文件名")

    filename = file.filename
    suffix = Path(filename).suffix.lower()
    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="文件内容为空")

    logger.info(
        "import_file start filename=%s suffix=%s size_bytes=%d",
        filename,
        suffix,
        len(data),
    )

    tables: list[LoadedTable]
    if suffix in {".xlsx", ".xls"}:
        try:
            tables = _load_tables_from_excel_stream(data)
        except Exception as exc:  # pragma: no cover - 细节错误交给调用方展示
            raise HTTPException(status_code=400, detail=f"解析 Excel 失败: {exc}") from exc
    elif suffix == ".csv":
        try:
            table_name = Path(filename).stem or "Sheet1"
            tables = _load_tables_from_csv(table_name, data)
        except Exception as exc:  # pragma: no cover
            raise HTTPException(status_code=400, detail=f"解析 CSV 失败: {exc}") from exc
    else:
        raise HTTPException(
            status_code=400,
            detail="仅支持 .xlsx, .xls, .csv 文件",
        )

    if not tables:
        raise HTTPException(
            status_code=400,
            detail="文件中未找到任何表数据",
        )

    tables_dict = {
        t.name: {
            "name": t.name,
            "rows": t.rows,
            "schema": [s.model_dump(mode="python") for s in t.table_schema],
        }
        for t in tables
    }
    state = project_store.create_project(tables_dict)
    logger.info(
        "import_file done project_id=%s tables=%d",
        state.id,
        len(tables),
    )
    return LoadSampleResponse(projectId=state.id, tables=tables)
