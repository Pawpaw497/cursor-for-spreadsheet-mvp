"""导出 Excel 相关路由。"""
import io
import time
from typing import Any

from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from pydantic import BaseModel, Field

from app.logging_config import get_logger

router = APIRouter(prefix="/api", tags=["export"])
log = get_logger("api.export")


class SchemaCol(BaseModel):
    key: str
    type: str  # "number" | "string" | "date"


class TableExport(BaseModel):
    model_config = {"populate_by_name": True}
    name: str
    rows: list[dict[str, Any]] = Field(default_factory=list)
    # 使用 schema_ + alias 避免与 BaseModel.schema 方法重名的警告
    schema_: list[SchemaCol] = Field(default_factory=list, alias="schema")


class ExportExcelRequest(BaseModel):
    tables: list[TableExport] = Field(min_length=1)


def _sanitize_sheet_name(name: str) -> str:
    """Excel 表名不能包含 : \\ / ? * [ ]，且长度 ≤ 31。"""
    invalid = (":", "\\", "/", "?", "*", "[", "]")
    out = name
    for c in invalid:
        out = out.replace(c, "_")
    return out[:31] if len(out) > 31 else out or "Sheet"


@router.post("/export-excel")
async def export_excel(req: ExportExcelRequest):
    """将当前项目多张表导出为一个 Excel 文件，每张表一个 sheet。"""
    t0 = time.perf_counter()
    shape = ";".join(
        f"{t.name}:r{len(t.rows)}c{len(t.schema_)}" for t in req.tables
    )
    log.info("export_excel start tables=%d shape=%s", len(req.tables), shape)
    wb = Workbook()
    # 删除默认创建的 sheet，后面按顺序创建
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    for t in req.tables:
        sheet_name = _sanitize_sheet_name(t.name)
        ws = wb.create_sheet(title=sheet_name)
        keys = [c.key for c in t.schema_] if t.schema_ else (
            list(t.rows[0].keys()) if t.rows else [])
        # 表头
        for col_idx, key in enumerate(keys, start=1):
            ws.cell(row=1, column=col_idx, value=key)
        # 数据行
        for row_idx, row in enumerate(t.rows, start=2):
            for col_idx, key in enumerate(keys, start=1):
                val = row.get(key)
                ws.cell(row=row_idx, column=col_idx, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    log.info(
        "export_excel done sheets=%d elapsed_ms=%.2f",
        len(req.tables),
        (time.perf_counter() - t0) * 1000,
    )
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=project.xlsx"},
    )
